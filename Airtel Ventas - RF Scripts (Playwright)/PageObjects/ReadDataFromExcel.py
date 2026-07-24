import datetime
import numpy as np
import openpyxl
import pandas as pd
import csv
import os

def fetch_no_rows(filename, sheetname):
    wk = openpyxl.load_workbook(filename)
    sh = wk[sheetname]
    return sh.max_row


def dummy():
    wk = 'hello'
    return wk


def fetch_from_exceldata(filename, sheetname, testcase_id, testdata_id):
    data = pd.read_excel(filename, sheet_name=sheetname)
    data.index = data['TestCaseId']
    dataframe = data.loc[[testcase_id]]
    dataframe.index = dataframe['TestDataId']
    test = dataframe.loc[[testdata_id]]
    return test


def fetch_cell_data(filename, sheetname, row, cell):
    wk = openpyxl.load_workbook(filename)
    sh = wk[sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value


def getdata(data, tag):
    dataset = data
    # pd.set_option('future.no_silent_downcasting', True)
    dataset = dataset.fillna('')
    if type(dataset[tag].iloc[0]) != str and dataset[tag].iloc[0] != 'NaN':
        dataset[tag] = dataset[tag].astype(np.int64)
    else:
        dataset[tag] = dataset[tag].iloc[0]

    return dataset[tag].values[0]


def get_value1(data, index):
    dataset = data
    return dataset[index].values[0]


def fetchrow(filename, sheetname, row):
    data = pd.read_excel(filename, sheet_name=sheetname)
    row = int(row)
    row_data = data.loc[[row]]
    return row_data


def getlistlen(list):
    newlist = [i for i in list.split(',')]
    return len(newlist)


def getlistdata(list, index):
    listval = []
    for i in list.split(','):
        listval.append(i)
    return listval[int(index)]


def getDateFromExcel(Data, Tag):
    dataSet = Data
    dataSet = dataSet.fillna('')
    if type(dataSet[Tag][0]) != str and dataSet[Tag][0] != 'NaN' and format(dataSet[Tag][0]) != 'NaT':
        if isinstance(dataSet[Tag][0], datetime.date):
            dataSet[Tag] = datetime.datetime.strptime(format(dataSet[Tag][0]), "%Y-%m-%d %H:%M:%S").date()
        else:
            dataSet[Tag] = dataSet[Tag].astype(np.int64)
    else:
        if format(dataSet[Tag][0]) == 'NaN' or format(dataSet[Tag][0]) == 'NaT':
            dataSet[Tag] = ''
        else:
            dataSet[Tag] = dataSet[Tag][0]
    return dataSet[Tag].values[0]


def fetch_from_excel_with_no_of_Entries(filename, sheetname, testcase_id, testdata_id):
    try:
        # Read the Excel file
        data = pd.read_excel(filename, sheet_name=sheetname)

        # Ensure required columns exist
        if 'TestCaseId' not in data.columns or 'TestDataId' not in data.columns:
            raise ValueError("Missing required columns 'TestCaseId' or 'TestDataId' in Excel sheet.")

        # Filter data based on TestCaseId & TestDataId
        filtered_data = data[(data['TestCaseId'] == testcase_id) & (data['TestDataId'] == testdata_id)]

        # **Check if DataFrame is empty**
        if filtered_data.empty:
            print(f"No data found for TestCaseId: {testcase_id} and TestDataId: {testdata_id}")
            return [], 0  # Return empty list and count as 0

        # Fill NaN values with empty strings (prevents errors)
        filtered_data = filtered_data.fillna("")

        # Count total number of entries
        total_entries = len(filtered_data)

        # Convert DataFrame to a list of dictionaries
        extracted_data = filtered_data.to_dict(orient='records')

        return extracted_data, total_entries  # Return extracted data and count

    except Exception as e:
        print(f"Error reading Excel: {e}")
        return [], 0  # Return empty data on error


def count_entries(filename, sheetname, testcase_id):
    try:
        # Read the Excel sheet
        data = pd.read_excel(filename, sheet_name=sheetname)

        # Ensure required columns exist
        if 'TestCaseId' not in data.columns:
            raise ValueError("Missing required column 'TestCaseId' in Excel sheet.")

        # Count matching rows for the given TestCaseId
        total_entries = len(data[data['TestCaseId'] == testcase_id])

        return total_entries  # Return count

    except Exception as e:
        print(f"Error reading Excel: {e}")
        return 0  # Return 0 on error


# def modify_csv_cell(folder, filename, columnname, row_index, value):
#     # Construct the file path
#     filepath = os.path.join(folder, filename)
#     print(f"File path: {filepath}")
#
#     # Read CSV into DataFrame, ensuring that the column is treated as a string
#     df = pd.read_csv(filepath, dtype={columnname: str})
#
#     # Ensure row index is within range
#     if row_index >= len(df):
#         raise ValueError(f"Row index {row_index} is out of range.")
#
#     # Check if column exists
#     if columnname not in df.columns:
#         raise ValueError(f"Column '{columnname}' not found in the CSV file.")
#
#     # Modify the specified cell (ensure value is treated as a string)
#     df.at[row_index, columnname] = str(value)
#
#     # Save back to CSV
#     df.to_csv(filepath, index=False)
#     print('File Edited and saved successfully')


import os

def modify_csv_cell(folder, filename, columnname, row_index, value):
    filepath = os.path.join(folder, filename)
    print(f"Editing file: {filepath}")

    # Read all lines from file
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    if not lines:
        raise ValueError("The CSV file is empty.")

    # Split header and validate column
    header = lines[0].strip().split(',')
    if columnname not in header:
        raise ValueError(f"Column '{columnname}' not found in header.")

    column_index = header.index(columnname)

    # Validate row index (skip header)
    if row_index + 1 >= len(lines):
        raise ValueError(f"Row index {row_index} is out of range.")

    # Split the target line and modify value
    row_data = lines[row_index + 1].strip().split(',')
    print(f"Original value: {row_data[column_index]}")
    row_data[column_index] = str(value)

    # Replace the line with updated data
    lines[row_index + 1] = ','.join(row_data) + '\n'

    # Write updated content back to file
    with open(filepath, 'w', encoding='utf-8') as f:
        f.writelines(lines)

    print("File edited and saved successfully.")



import os
from openpyxl import load_workbook
from copy import copy

def modify_excel_cell(folder, filename, columnname, row_index, value):
    filepath = os.path.join(folder, filename)
    print(filepath)

    wb = load_workbook(filepath)
    ws = wb.active  # Keeps original sheet name

    # Find column index by header row (row 1)
    col_index = None
    for cell in ws[1]:
        if str(cell.value).strip() == str(columnname).strip():
            col_index = cell.column
            break
    if col_index is None:
        raise ValueError(f"Column '{columnname}' not found in the Excel file.")

    excel_row = row_index + 2  # +1 for header, +1 because 0-based index

    # Get existing cell
    old_cell = ws.cell(row=excel_row, column=col_index)

    # Update value
    old_cell.value = value  # Editing the same cell preserves formatting automatically

    wb.save(filepath)
    print('File Edited and saved Successfully')