"""
Generate QA Infinity test case upload Excel for POS Sales & Return module.
Columns match the seed-tc-template.xlsx expected by qa-infinity's parse-seed endpoint.

Run:  python generate_qa_infinity_tc_upload.py
Output: QA_Infinity_TC_Upload_POS_Sales.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Use Case",
    "Title",
    "Objective",
    "Priority",
    "Test Type",
    "Pre-conditions / Dependencies",
    "Test Steps",
    "Test Data",
    "Expected Result",
    "Actual Result",
    "Notes",
]

# ---------------------------------------------------------------------------
# Test cases data
# ---------------------------------------------------------------------------

TEST_CASES = [
    # ---- Physical Product ----
    {
        "use_case": "POS Sales",
        "title": "TC01 - Create POS Sale order - Physical product - New customer",
        "objective": "Validate that a Cashier can create a POS sale order for a physical product with a brand-new customer and complete payment via Cash.",
        "priority": "CRITICAL",
        "type": "UI",
        "pre_conditions": "Cashier user is active. At least one physical product with stock is available in the cash register. Test data row TC01/TD01 exists in the POSsales sheet.",
        "steps": (
            "1. Login as Cashier.\n"
            "2. Navigate to POS Sales > POS Sales sub-menu.\n"
            "3. Click Create Order and select 'Enter Name' to open customer search.\n"
            "4. Type '@' to trigger the Add New Customer option.\n"
            "5. Fill First Name (generated), Contact Number, Email, Invoice Type, Address Line 1 & 2.\n"
            "6. Submit the new customer form.\n"
            "7. Scan the serial number of the physical product.\n"
            "8. Verify payment options (Cash, Cheque, Credit Card, Bank Transfer, Airtel Money) are visible.\n"
            "9. Select Cash, enter amount equal to item price, select bank code, click Add.\n"
            "10. Click Submit > Process Order > Confirm.\n"
            "11. Capture Order ID and Customer ID from the receipt.\n"
            "12. Logout and re-login as Cashier.\n"
            "13. Navigate to POS Sales list and search by Order ID.\n"
            "14. Verify order status = COMPLETED, invoice status = GENERATED.\n"
            "15. Open Track Order panel and verify all suborders are Completed."
        ),
        "test_data": "Login: Airtel_Cashier_Login / Airtel_Cashier_TD_Login\nPOSsales sheet: TC01 / TD01 (InvoiceType, CashAmount, BankCode)",
        "expected": (
            "- New customer is created with auto-generated name, contact, email.\n"
            "- POS sale order is created with status COMPLETED.\n"
            "- Invoice status = GENERATED.\n"
            "- Customer receipt shows correct name, email, contact, total price.\n"
            "- Suborders (erp-invoice-posting, RA-integration, esb-inventory-posting) all = Completed."
        ),
    },
    {
        "use_case": "POS Sales",
        "title": "TC02 - Create POS Sale order - Physical product - Existing customer",
        "objective": "Validate that a Cashier can search and reuse an existing customer for a second POS sale order.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "TC01 must have run successfully so a customer record exists. Asset available in stock.",
        "steps": (
            "1. Login as Cashier and complete a new-customer POS sale (same as TC01).\n"
            "2. Logout and re-login.\n"
            "3. Navigate to POS Sales.\n"
            "4. Click Create Order > Enter Name, type the previously generated customer name.\n"
            "5. Select the customer from search results.\n"
            "6. Verify customer name and contact number are pre-filled correctly.\n"
            "7. Scan a valid product serial number (enter an invalid one first to trigger error).\n"
            "8. Pay via Cash and submit the order.\n"
            "9. Verify order status and suborders."
        ),
        "test_data": "Customer name / contact stored from TC01 run. POSsales sheet: TC01 / TD01.",
        "expected": (
            "- Existing customer is found via search.\n"
            "- Second POS sale order is created and COMPLETED.\n"
            "- Invalid serial number scan shows error 'Product not found'.\n"
            "- Valid serial number scan shows correct serial, quantity and price."
        ),
    },
    {
        "use_case": "POS Sales",
        "title": "TC03 - Create POS Sale order - Physical product - Partner",
        "objective": "Validate that a Cashier can create a POS sale order selecting a partner (distributor) as buyer.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "Partner name, number and email available in POSsales TC01/TD01 test data. Asset in stock.",
        "steps": (
            "1. Login as Cashier.\n"
            "2. Navigate to cash register, store asset serial from stock view.\n"
            "3. Navigate to POS Sales > Create Order.\n"
            "4. Click 'Partner' toggle.\n"
            "5. Search and select the partner from overlay.\n"
            "6. Verify partner name and number.\n"
            "7. Scan product serial number.\n"
            "8. Pay via Cash and submit.\n"
            "9. Verify order in POS Sales list."
        ),
        "test_data": "POSsales sheet: TC01 / TD01 (Partner, PartnerNumber, PartnerMailID).",
        "expected": "Order created with partner as buyer. Status = COMPLETED. Invoice generated.",
    },
    {
        "use_case": "POS Sales",
        "title": "TC04 - Create POS Sale order - Physical product - Multiple products",
        "objective": "Validate that two separate serial numbers can be added to one POS cart, with ability to remove a mistakenly scanned third serial.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "Existing customer record available. Three physical product assets in stock.",
        "steps": (
            "1. Login as Cashier, retrieve assets from stock view.\n"
            "2. Navigate to POS Sales, search existing customer.\n"
            "3. Scan first serial number.\n"
            "4. Scan second serial number.\n"
            "5. Scan a third serial number, then remove it using the remove icon and confirm.\n"
            "6. Verify only first and second serials appear in cart.\n"
            "7. Pay via Cash (full cart amount) and submit.\n"
            "8. Verify both serial numbers appear in the completed order."
        ),
        "test_data": "Three AssetValues stored from stock view. POSsales sheet: TC01 / TD01.",
        "expected": "Cart shows exactly 2 products. Third serial removed successfully. Order COMPLETED with Qty=2.",
    },
    {
        "use_case": "POS Sales",
        "title": "TC05 - Create POS Sale order - Pay using multiple payment modes (Cash + Cheque)",
        "objective": "Validate that the cart amount can be split across two payment modes (50% Cash, 50% Cheque).",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "Existing customer available. Asset in stock. Bank Name and Bank Code configured in test data.",
        "steps": (
            "1. Login as Cashier, retrieve asset, navigate to POS Sales, search existing customer.\n"
            "2. Scan product serial number.\n"
            "3. Select Cash payment option, enter 50% of item price, select bank code, click Add.\n"
            "4. Verify Cash payment line shows correct amount and method.\n"
            "5. Select Cheque payment option, set date (today), select Bank Name and Bank Code.\n"
            "6. Enter payee name (customer name), enter remaining 50% as cheque amount, enter cheque number.\n"
            "7. Click Add for cheque.\n"
            "8. Verify both payment lines are shown.\n"
            "9. Submit and confirm order.\n"
            "10. Verify order in list with both payment methods."
        ),
        "test_data": "POSsales sheet: TC01 / TD01 (BankName, BankCode). Split amounts derived from item price.",
        "expected": "Order COMPLETED. Payment details show Cash amount and Cheque amount summing to total.",
    },
    {
        "use_case": "POS Sales",
        "title": "TC06 - Create POS Sale order - Pay using Airtel Money",
        "objective": "Validate POS sale payment via Airtel Money (MPESA-style mobile money), including transaction validation.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "Existing customer. Asset in stock. Valid Airtel Money transaction ID available.",
        "steps": (
            "1. Login as Cashier, retrieve asset, search existing customer.\n"
            "2. Scan product serial.\n"
            "3. Select Airtel Money/Mobile Money payment option.\n"
            "4. Enter Transaction ID and transaction amount.\n"
            "5. Click Validate; verify message 'Transaction validated successfully.'.\n"
            "6. Click Add.\n"
            "7. If Airtel Money amount < total, pay remainder via Cash.\n"
            "8. Submit and confirm order.\n"
            "9. Verify order in list with Airtel Money payment detail."
        ),
        "test_data": "POSsales sheet: TC01 / TD01 (TransactionID, AirtelMoneyAmount).",
        "expected": "Transaction validated. Order COMPLETED. Payment details show Airtel Money amount.",
    },
    {
        "use_case": "POS Sales",
        "title": "TC07 - Create POS Sale order - Pay more amount than cart total",
        "objective": "Validate that paying more than the cart amount triggers the change calculation in the cash register.",
        "priority": "MEDIUM",
        "type": "UI",
        "pre_conditions": "Existing customer and asset in stock. CashAmount in test data is greater than item price.",
        "steps": (
            "1. Login as Cashier, search existing customer, scan product.\n"
            "2. Select Cash payment, enter an amount higher than item price.\n"
            "3. Add payment, submit and process order.\n"
            "4. Navigate to cash register and verify change value is correctly calculated."
        ),
        "test_data": "POSsales sheet: TC01 / TD01 (CashAmount > item price).",
        "expected": "Order created. Cash register displays correct change amount (excess paid - item price).",
    },
    {
        "use_case": "POS Sales",
        "title": "TC08 - Check POS Sale order - Park customer",
        "objective": "Validate the Park Order feature — saves an incomplete order for later without processing.",
        "priority": "MEDIUM",
        "type": "UI",
        "pre_conditions": "Customer and asset available. Park feature is enabled for the role.",
        "steps": (
            "1. Login as Cashier, create sale order with product in cart.\n"
            "2. Click Park Order button instead of Submit.\n"
            "3. Verify order is saved in parked state.\n"
            "4. Retrieve the parked order and resume it.\n"
            "5. Complete payment and submit."
        ),
        "test_data": "POSsales sheet: TC01 / TD01.",
        "expected": "Order is parked and retrievable. After resuming, order completes with status COMPLETED.",
    },
    {
        "use_case": "POS Sales",
        "title": "TC09 - Verify Invoice and Payment Receipt after POS Sale order",
        "objective": "Validate that both the invoice PDF and payment receipt PDF can be downloaded after order completion.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "A completed POS sale order exists (run TC01 first). Download directory set.",
        "steps": (
            "1. Login as Cashier and navigate to POS Sales list.\n"
            "2. Search and open the completed order from TC01.\n"
            "3. Click Actions > Download Invoice.\n"
            "4. Verify the PDF file is downloaded (non-temp, non-empty).\n"
            "5. Click Actions > Download Payment Receipt.\n"
            "6. Verify the PDF file is downloaded."
        ),
        "test_data": "Order ID from TC01 run. Download directory path from suite setup.",
        "expected": "Invoice PDF downloaded successfully. Payment Receipt PDF downloaded successfully. No temp (.crdownload) files.",
    },
    {
        "use_case": "POS Sales - Returns",
        "title": "TC10 - Create POS Sale return - Using partner",
        "objective": "Validate return order creation for a sale originally made to a partner.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "TC03 (Partner sale) must have completed successfully.",
        "steps": (
            "1. Login as Cashier, navigate to POS Sales list.\n"
            "2. Search and open the order from TC03 (partner sale).\n"
            "3. Click Actions > Return Item.\n"
            "4. Select return type and confirm.\n"
            "5. Submit return order.\n"
            "6. Search the return order by Order ID.\n"
            "7. Verify return order type = PRV_ORD, status = COMPLETED."
        ),
        "test_data": "Order ID from TC03. Return type from POSsales test data.",
        "expected": "Return order created with type PRV_ORD. Invoice generated. Status = COMPLETED.",
    },
    {
        "use_case": "POS Sales - Returns",
        "title": "TC11 - Create POS Sale return - Using customer",
        "objective": "Validate return order creation for a sale originally made to an end customer.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "TC01 or TC02 (customer sale) must have completed successfully.",
        "steps": (
            "1. Login as Cashier, navigate to POS Sales list.\n"
            "2. Search and open the completed customer sale order.\n"
            "3. Click Actions > Return Item, confirm.\n"
            "4. Verify return order in list: type PRV_ORD, status COMPLETED."
        ),
        "test_data": "Order ID from TC01/TC02 run.",
        "expected": "Return order PRV_ORD created. Status = COMPLETED. Invoice generated.",
    },
    {
        "use_case": "POS Sales - Returns",
        "title": "TC12 - Create POS Sale return - Product condition Damaged",
        "objective": "Validate return where the product condition is marked as Damaged.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "A completed POS sale exists.",
        "steps": (
            "1. Navigate to completed POS sale order.\n"
            "2. Click Actions > Return Item.\n"
            "3. Select Return Type = Damaged.\n"
            "4. Confirm and submit.\n"
            "5. Verify return order with Damaged condition."
        ),
        "test_data": "POSsales sheet: return type = Damaged.",
        "expected": "Return order created with product condition = Damaged. Status COMPLETED.",
    },
    {
        "use_case": "POS Sales - Returns",
        "title": "TC13 - Create POS Sale return - Multiple assets",
        "objective": "Validate return of an order that contained multiple product assets.",
        "priority": "MEDIUM",
        "type": "UI",
        "pre_conditions": "TC04 (multi-product sale) must have completed successfully.",
        "steps": (
            "1. Navigate to TC04 order in POS Sales list.\n"
            "2. Click Actions > Return Item.\n"
            "3. All assets should be listed; confirm return.\n"
            "4. Verify all assets are returned and order status is COMPLETED."
        ),
        "test_data": "Order ID from TC04.",
        "expected": "All multiple assets are returned. Return order type PRV_ORD. Status COMPLETED.",
    },
    {
        "use_case": "POS Sales - Returns",
        "title": "TC14 - Verify invoice generation after return order",
        "objective": "Validate that invoice is generated after a POS return order is completed.",
        "priority": "MEDIUM",
        "type": "UI",
        "pre_conditions": "A completed return order exists.",
        "steps": (
            "1. Navigate to a completed POS return order.\n"
            "2. Click Actions.\n"
            "3. Verify Download Invoice option is present (for return) and download succeeds."
        ),
        "test_data": "Return Order ID from TC10 or TC11.",
        "expected": "Invoice PDF downloads successfully for the return order.",
    },
    {
        "use_case": "POS Sales - Replace",
        "title": "TC15 - Create POS Sale replace - Using partner",
        "objective": "Validate replacement order creation for a partner sale.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "TC03 (partner sale) completed. Replacement asset available in stock.",
        "steps": (
            "1. Navigate to TC03 partner order in POS Sales list.\n"
            "2. Click Actions > Replace Item.\n"
            "3. Scan replacement asset serial number.\n"
            "4. Confirm replacement.\n"
            "5. Verify replacement order type PRP_ORD in order list."
        ),
        "test_data": "Order ID from TC03. Replacement asset serial from stock.",
        "expected": "Replace order PRP_ORD created. Original asset deactivated. Replacement serial assigned.",
    },
    {
        "use_case": "POS Sales - Replace",
        "title": "TC16 - Create POS Sale replace - Using customer",
        "objective": "Validate replacement order for an end-customer sale.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "TC01 (customer sale) completed. Replacement asset in stock.",
        "steps": (
            "1. Navigate to TC01 customer order in POS Sales list.\n"
            "2. Click Actions > Replace Item.\n"
            "3. Scan replacement serial number.\n"
            "4. Confirm replacement.\n"
            "5. Verify replace order PRP_ORD in list."
        ),
        "test_data": "Order ID from TC01. Replacement serial from stock.",
        "expected": "Replace order PRP_ORD created successfully.",
    },
    {
        "use_case": "POS Sales - EVD",
        "title": "TC17 - Create POS Sale EVD - Virtual product - New customer",
        "objective": "Validate EVD (Electronic Virtual Delivery / recharge) POS sale for a new customer, including cash register balance verification.",
        "priority": "CRITICAL",
        "type": "UI",
        "pre_conditions": "Cashier has sufficient cash register balance. EVD test data (MSISDN, EVDAmount) in test data.",
        "steps": (
            "1. Login as Cashier.\n"
            "2. Navigate to Cash Register and record opening balance.\n"
            "3. Navigate to POS Sales > POS Sales EVD.\n"
            "4. Click Create Order, search '@', add new customer.\n"
            "5. Fill customer details (name, contact, email, invoice type, address).\n"
            "6. Enter MSISDN and EVD/recharge amount.\n"
            "7. Verify payment options.\n"
            "8. Pay by Cash and click Recharge.\n"
            "9. Submit and verify EVD order receipt (Customer ID, name, email, amount, order ID).\n"
            "10. Navigate back to POS Sales EVD list and search the order.\n"
            "11. Verify status COMPLETED, invoice GENERATED, customer type END_CUSTOMER.\n"
            "12. Navigate to Cash Register and verify closing balance = opening - EVD amount."
        ),
        "test_data": "POSsales sheet: TC01 / TD01 (MSISDN, EVDAmount, InvoiceType).",
        "expected": (
            "EVD sale order created and COMPLETED.\n"
            "Customer receipt shows correct name, email, contact and total amount.\n"
            "Cash register balance decreased by EVD amount.\n"
            "Invoice status = GENERATED."
        ),
    },
    {
        "use_case": "POS Sales - EVD",
        "title": "TC18 - Create POS Sale EVD - Multiple payment modes",
        "objective": "Validate EVD sale paid by split across Cash and a second payment mode.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "Existing customer. EVD test data available.",
        "steps": (
            "1. Login as Cashier.\n"
            "2. Navigate to POS Sales EVD, search existing customer.\n"
            "3. Enter MSISDN and EVD amount.\n"
            "4. Pay 50% via Cash, 50% via second mode.\n"
            "5. Click Recharge and submit.\n"
            "6. Verify order in EVD list."
        ),
        "test_data": "POSsales sheet: TC01 / TD01 (Split EVDAmount).",
        "expected": "EVD order COMPLETED with two payment line items summing to total EVD amount.",
    },
    {
        "use_case": "POS Sales - EVD",
        "title": "TC19 - Create POS Sale EVD - Pay using Airtel Money",
        "objective": "Validate EVD sale paid entirely via Airtel Money transaction.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "Existing customer. Valid Airtel Money transaction ID in test data.",
        "steps": (
            "1. Login as Cashier.\n"
            "2. Navigate to POS Sales EVD, search existing customer.\n"
            "3. Enter MSISDN and EVD amount.\n"
            "4. Select Airtel Money, enter Transaction ID.\n"
            "5. Click Add, then Recharge and submit.\n"
            "6. Verify EVD order in list."
        ),
        "test_data": "POSsales sheet: TC01 / TD01 (TransactionID, EVDAmount).",
        "expected": "EVD order COMPLETED. Payment method = Airtel Money.",
    },
    {
        "use_case": "POS Sales - EVD",
        "title": "TC20 - Verify Invoice and Payment Receipt after POS Sale EVD",
        "objective": "Validate that invoice and payment receipt PDFs can be downloaded after an EVD sale.",
        "priority": "MEDIUM",
        "type": "UI",
        "pre_conditions": "TC17 EVD order completed. Download directory set.",
        "steps": (
            "1. Navigate to the completed EVD order from TC17.\n"
            "2. Click Actions > Download Invoice, verify PDF downloaded.\n"
            "3. Click Actions > Download Payment Receipt, verify PDF downloaded."
        ),
        "test_data": "EVD Order ID from TC17.",
        "expected": "Invoice and Payment Receipt PDFs downloaded. No temp files.",
    },
    {
        "use_case": "POS Sales - EVD",
        "title": "TC21 - Create POS EVD order - Pay more than topup amount",
        "objective": "Validate that paying more than the EVD recharge amount triggers correct change/excess calculation.",
        "priority": "MEDIUM",
        "type": "UI",
        "pre_conditions": "Existing customer. CashAmount > EVDAmount in test data.",
        "steps": (
            "1. Navigate to POS Sales EVD, search existing customer.\n"
            "2. Enter MSISDN and EVD amount.\n"
            "3. Select Cash, enter amount greater than EVD amount.\n"
            "4. Submit and verify overpayment is handled (change displayed in cash register)."
        ),
        "test_data": "POSsales sheet: TC01 / TD01 (CashAmount > EVDAmount).",
        "expected": "EVD order COMPLETED. Cash register shows correct change value.",
    },
    {
        "use_case": "POS Sales - EVD Bulk",
        "title": "TC22 - Create EVD Bulk Order with proper file",
        "objective": "Validate that uploading a correctly formatted CSV file processes all recharge orders successfully.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "POS_EVD_Bulk.csv file available with correct MSISDN and amount columns.",
        "steps": (
            "1. Login as Cashier.\n"
            "2. Navigate to POS Sales > POS Sales EVD.\n"
            "3. Click Bulk ERCV option.\n"
            "4. Upload POS_EVD_Bulk.csv.\n"
            "5. Pay via Cash.\n"
            "6. Click Recharge and submit.\n"
            "7. Verify all bulk orders have COMPLETED status."
        ),
        "test_data": "FilePath: POS_EVD_Bulk.csv from test data config.",
        "expected": "All MSISDN records recharged. Order COMPLETED. Invoice generated.",
    },
    {
        "use_case": "POS Sales - EVD Bulk",
        "title": "TC23 - Create EVD Bulk Order with duplicate MSISDN in file",
        "objective": "Validate negative scenario where CSV contains duplicate MSISDN entries.",
        "priority": "HIGH",
        "type": "UI",
        "pre_conditions": "POS_EVD_Bulk_Duplicate_MSISDN.csv available.",
        "steps": (
            "1. Login as Cashier.\n"
            "2. Navigate to POS Sales EVD > Bulk ERCV.\n"
            "3. Upload the duplicate MSISDN CSV file.\n"
            "4. Pay via Cash and attempt to process.\n"
            "5. Verify error message is displayed."
        ),
        "test_data": "FilePath: POS_EVD_Bulk_Duplicate_MSISDN.csv.",
        "expected": "Error message displayed: 'File has duplicate records'. Order is not processed.",
    },
    {
        "use_case": "POS Sales - EVD Bulk",
        "title": "TC24 - Create EVD Bulk Order with empty file",
        "objective": "Validate negative scenario where an empty CSV is uploaded for bulk EVD.",
        "priority": "MEDIUM",
        "type": "UI",
        "pre_conditions": "POS_EVD_Bulk_Empty_File.csv available.",
        "steps": (
            "1. Login as Cashier.\n"
            "2. Navigate to POS Sales EVD > Bulk ERCV.\n"
            "3. Upload the empty CSV file.\n"
            "4. Pay via Cash and attempt to process.\n"
            "5. Verify the error message."
        ),
        "test_data": "FilePath: POS_EVD_Bulk_Empty_File.csv.",
        "expected": "Error message: 'Transfer amount provided is not matching with amount from file'. Order not created.",
    },
]


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def make_header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="1F4E79"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def make_data_style(row_idx):
    bg = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
    return {
        "fill": PatternFill("solid", fgColor=bg),
        "alignment": Alignment(vertical="top", wrap_text=True),
        "border": Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        ),
    }


def apply_style(cell, style_dict):
    for attr, val in style_dict.items():
        setattr(cell, attr, val)


def generate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POS Sales Test Cases"
    ws.freeze_panes = "A2"

    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_style(cell, make_header_style())

    # Write test case rows
    for row_idx, tc in enumerate(TEST_CASES, start=2):
        row_values = [
            tc["use_case"],
            tc["title"],
            tc["objective"],
            tc["priority"],
            tc["type"],
            tc["pre_conditions"],
            tc["steps"],
            tc["test_data"],
            tc["expected"],
            "",   # Actual Result (blank — filled after execution)
            "",   # Notes
        ]
        style = make_data_style(row_idx)
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_style(cell, style)

    # Column widths
    col_widths = {
        1: 22,   # Use Case
        2: 55,   # Title
        3: 50,   # Objective
        4: 12,   # Priority
        5: 10,   # Type
        6: 45,   # Pre-conditions
        7: 80,   # Test Steps
        8: 45,   # Test Data
        9: 60,   # Expected Result
        10: 25,  # Actual Result
        11: 25,  # Notes
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Row height — data rows taller to fit wrapped text
    ws.row_dimensions[1].height = 28
    for row_idx in range(2, len(TEST_CASES) + 2):
        ws.row_dimensions[row_idx].height = 120

    output_path = "QA_Infinity_TC_Upload_POS_Sales.xlsx"
    wb.save(output_path)
    print(f"Generated: {output_path}  ({len(TEST_CASES)} test cases)")


if __name__ == "__main__":
    generate()
